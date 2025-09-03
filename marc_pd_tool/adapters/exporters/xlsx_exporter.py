# marc_pd_tool/adapters/exporters/xlsx_exporter.py

"""Standard XLSX export functionality that reads from JSON data"""

# Standard library imports

# Third party imports
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Local imports
from marc_pd_tool.adapters.exporters.base_exporter import BaseJSONExporter
from marc_pd_tool.core.types.json import JSONDict
from marc_pd_tool.core.types.json import JSONList
from marc_pd_tool.core.types.json import JSONType


class XLSXExporter(BaseJSONExporter):
    """Export standard XLSX files from JSON data

    Creates Excel files with separate tabs for each copyright status,
    showing records in a tabular format with match details.
    """

    __slots__ = ("country_codes",)

    def __init__(self, json_path: str, output_path: str, single_file: bool = False):
        """Initialize the exporter with JSON data and country code mappings"""
        super().__init__(json_path, output_path, single_file)

        # Country code mappings (MARC country codes to display names)
        self.country_codes = {
            "abc": "Alberta",
            "ae": "Algeria",
            "ag": "Argentina",
            "ai": "Anguilla",
            "at": "Australia",
            "au": "Austria",
            "be": "Belgium",
            "bl": "Brazil",
            "bu": "Bulgaria",
            "bw": "Belarus",
            "cc": "China",
            "ch": "China (Republic)",
            "ck": "Colombia",
            "cl": "Chile",
            "cs": "Czechoslovakia",
            "cu": "Cuba",
            "cy": "Cyprus",
            "dk": "Denmark",
            "enk": "United Kingdom",
            "es": "El Salvador",
            "fr": "France",
            "ge": "Germany (East)",
            "gh": "Ghana",
            "gr": "Greece",
            "gs": "Georgia (Republic)",
            "gt": "Guatemala",
            "gw": "Germany",
            "hk": "Hong Kong",
            "ht": "Haiti",
            "hu": "Hungary",
            "ie": "Ireland",
            "ii": "India",
            "iq": "Iraq",
            "ir": "Iran",
            "is": "Israel",
            "it": "Italy",
            "iv": "Ivory Coast",
            "ja": "Japan",
            "jo": "Jordan",
            "ko": "Korea (South)",
            "le": "Lebanon",
            "li": "Lithuania",
            "lu": "Luxembourg",
            "lv": "Latvia",
            "mk": "Macedonia",
            "mm": "Malta",
            "mr": "Morocco",
            "mx": "Mexico",
            "ne": "Netherlands",
            "no": "Norway",
            "nr": "Nigeria",
            "onc": "Ontario",
            "pe": "Peru",
            "pl": "Poland",
            "po": "Portugal",
            "pr": "Puerto Rico",
            "quc": "Quebec",
            "rm": "Romania",
            "ru": "Russia",
            "rur": "Russia (Federation)",
            "sa": "South Africa",
            "si": "Singapore",
            "sp": "Spain",
            "stk": "Scotland",
            "sw": "Sweden",
            "sy": "Syria",
            "sz": "Switzerland",
            "ta": "Tajikistan",
            "th": "Thailand",
            "ti": "Tunisia",
            "tu": "Turkey",
            "ua": "Egypt",
            "uik": "United Kingdom (Misc.)",
            "un": "Soviet Union",
            "unr": "Soviet Union (Regions)",
            "us": "United States",
            "uz": "Uzbekistan",
            "vc": "Vatican City",
            "ve": "Venezuela",
            "vm": "Vietnam",
            "vn": "Vietnam (North)",
            "vp": "Various places",
            "wb": "West Berlin",
            "wiu": "Wisconsin",
            "wlk": "Wales",
            "xr": "Czech Republic",
            "xx": "Unknown",
            "xxc": "Canada",
            "xxk": "United Kingdom",
            "yu": "Yugoslavia",
        }

    # Header styling
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data styling
    DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Score coloring
    HIGH_SCORE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    MEDIUM_SCORE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    LOW_SCORE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    def _get_default_status_color(self, status: str) -> str:
        """Get a default color for a status based on its content patterns

        Args:
            status: The status string (e.g., "US_PRE_1929", "FOREIGN_RENEWED_FRA")

        Returns:
            A hex color code
        """
        status_upper = status.upper()

        # US statuses - various shades of green for PD, red for copyright
        if status_upper.startswith("US_"):
            if "PRE_" in status_upper:  # US_PRE_1929, etc.
                return "E8F5E9"  # Light green - oldest PD
            elif "REGISTERED_NOT_RENEWED" in status_upper:
                return "D5F4E6"  # Green - PD due to non-renewal
            elif "NO_RENEWAL" in status_upper or "REG_NO_RENEWAL" in status_upper:
                return "D4E6F1"  # Light blue-green - PD
            elif "RENEWED" in status_upper:
                return "FFCDD2"  # Light red - still in copyright
            elif "NO_MATCH" in status_upper or "NO_REG" in status_upper:
                return "E3F2FD"  # Light blue - unknown status
            else:
                return "FCF3CF"  # Light yellow - needs research

        # Foreign statuses - purples and pinks
        elif status_upper.startswith("FOREIGN_"):
            if "PRE_" in status_upper:
                return "E8DAEF"  # Light purple - likely PD
            elif "RENEWED" in status_upper:
                return "FADBD8"  # Light pink - may be in copyright
            elif "REGISTERED_NOT_RENEWED" in status_upper:
                return "F3E5F5"  # Very light purple
            else:
                return "FCF3CF"  # Light yellow - needs research

        # Unknown country - grays
        elif "UNKNOWN" in status_upper:
            if "RENEWED" in status_upper:
                return "FFCDD2"  # Light red - likely in copyright
            elif "REGISTERED_NOT_RENEWED" in status_upper:
                return "E0E0E0"  # Medium gray
            else:
                return "E5E7E9"  # Light gray

        # Out of data range - yellow
        elif "OUT_OF_DATA_RANGE" in status_upper:
            return "FFF9C4"  # Very light yellow

        # Default light gray for unrecognized patterns
        else:
            return "F5F5F5"

    def _format_status_as_sheet_name(self, status: str) -> str:
        """Format a status string as a valid Excel sheet name

        Args:
            status: The status string (e.g., "US_RENEWED", "FOREIGN_NO_MATCH_GBR")

        Returns:
            A sheet name suitable for Excel (max 31 chars)
        """
        # Common replacements to shorten and make readable
        replacements = {
            "REGISTERED_NOT_RENEWED": "Reg Not Renewed",
            "REGISTERED": "Reg",
            "FOREIGN": "For",
            "COUNTRY_UNKNOWN": "Unknown",
            "OUT_OF_DATA_RANGE": "Out Range",
            "NO_MATCH": "No Match",
            "RENEWED": "Renewed",
            "PRE_": "Pre ",
        }

        formatted = status
        for old, new in replacements.items():
            formatted = formatted.replace(old, new)

        # Replace underscores with spaces
        formatted = formatted.replace("_", " ")

        # Title case but keep country codes uppercase
        parts = formatted.split()
        formatted_parts = []
        for part in parts:
            # Keep 3-letter country codes uppercase
            if len(part) == 3 and part.isalpha() and part.isupper():
                formatted_parts.append(part)
            else:
                formatted_parts.append(part.title())
        formatted = " ".join(formatted_parts)

        # Excel sheet names have a hard limit of 31 characters
        if len(formatted) > 31:
            # Try to truncate intelligently at word boundary
            if len(formatted) > 28:
                formatted = formatted[:28] + "..."
            else:
                formatted = formatted[:31]

        return formatted

    def export(self) -> None:
        """Export records to XLSX file"""
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create summary sheet
        self._create_summary_sheet(wb)

        if self.single_file:
            # All records in one sheet
            records = self.records
            sorted_records = self.sort_by_quality(records)
            self._create_data_sheet(wb, "All Records", sorted_records)
        else:
            # Group by status and create sheets
            by_status = self.group_by_status()

            # Group statuses by category for organized tabs
            grouped = self._group_for_tabs(by_status)

            # Create tabs in organized order
            for tab_name, records in grouped:
                if records:
                    sorted_records = self.sort_by_quality(records)
                    self._create_data_sheet(wb, tab_name, sorted_records)

        # Save the workbook
        wb.save(self.output_path)

    def _group_for_tabs(self, by_status: dict[str, JSONList]) -> list[tuple[str, JSONList]]:
        """Group statuses into organized tabs

        Returns list of (tab_name, records) tuples in display order
        """
        tabs = []

        # Group by category
        us_records: dict[str, JSONList] = {}
        foreign_by_type: dict[str, JSONList] = {}
        unknown_records: dict[str, JSONList] = {}
        out_of_range_records: JSONList = []

        for status, records in by_status.items():
            if not records:
                continue

            if status.startswith("US_"):
                # US tabs
                status_type = status[3:]
                us_records[status_type] = records
            elif status.startswith("FOREIGN_"):
                # Parse foreign status type
                parts = status.split("_", 2)
                if len(parts) >= 3:
                    status_part = parts[1]
                    remaining = parts[2]

                    # Check if this is a multi-word status
                    if status_part == "NO" and remaining.startswith("MATCH_"):
                        status_type = "NO_MATCH"
                    elif status_part == "PRE" and remaining.startswith("1929_"):
                        status_type = "PRE_1929"
                    elif status_part == "REGISTERED" and remaining.startswith("NOT_RENEWED_"):
                        status_type = "REGISTERED_NOT_RENEWED"
                    else:
                        status_type = status_part

                    if status_type not in foreign_by_type:
                        foreign_by_type[status_type] = []
                    foreign_by_type[status_type].extend(records)
            elif status.startswith("COUNTRY_UNKNOWN"):
                status_type = status[16:] if len(status) > 16 else "UNKNOWN"
                unknown_records[status_type] = records
            elif status.startswith("OUT_OF_DATA_RANGE"):
                out_of_range_records.extend(records)

        # Add US tabs first
        for status_type in ["NO_MATCH", "PRE_1929", "REGISTERED_NOT_RENEWED", "RENEWED"]:
            if status_type in us_records:
                tab_name = f"US {self._format_status_type(status_type)}"
                tabs.append((tab_name, us_records[status_type]))

        # Add Foreign tabs
        for status_type in ["NO_MATCH", "PRE_1929", "REGISTERED_NOT_RENEWED", "RENEWED"]:
            if status_type in foreign_by_type:
                tab_name = f"Foreign {self._format_status_type(status_type)}"
                tabs.append((tab_name, foreign_by_type[status_type]))

        # Add Country Unknown tabs
        for status_type in ["NO_MATCH", "PRE_1929", "REGISTERED_NOT_RENEWED", "RENEWED"]:
            if status_type in unknown_records:
                tab_name = f"Unknown {self._format_status_type(status_type)}"
                tabs.append((tab_name, unknown_records[status_type]))

        # Add Out of Range tab if present
        if out_of_range_records:
            tabs.append(("Out of Data Range", out_of_range_records))

        return tabs

    def _format_status_type(self, status_type: str) -> str:
        """Format status type for display"""
        formats = {
            "NO_MATCH": "No Match",
            "PRE_1929": "Pre-1929",
            "REGISTERED_NOT_RENEWED": "Not Renewed",
            "RENEWED": "Renewed",
        }
        return formats.get(status_type, status_type.replace("_", " ").title())

    def _create_summary_sheet(self, wb: Workbook) -> None:
        """Create summary sheet with statistics"""
        ws = wb.create_sheet("Summary")

        # Title
        ws["A1"] = "MARC PD Tool Analysis Results"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:B1")

        # Processing info
        metadata = self.metadata
        ws["A3"] = "Processing Date:"
        ws["B3"] = metadata.get("processing_date", "Unknown")

        ws["A4"] = "Total Records:"
        ws["B4"] = metadata.get("total_records", 0)

        ws["A5"] = "Tool Version:"
        ws["B5"] = metadata.get("tool_version", "Unknown")

        # Processing parameters section
        ws["A7"] = "Processing Parameters:"
        ws["A7"].font = Font(bold=True)

        parameters = metadata.get("parameters", {})
        param_row = 8
        if isinstance(parameters, dict):
            param_info = [
                ("Title Threshold:", parameters.get("title_threshold", "Unknown")),
                ("Author Threshold:", parameters.get("author_threshold", "Unknown")),
                ("Publisher Threshold:", parameters.get("publisher_threshold", "Unknown")),
                ("Year Tolerance:", parameters.get("year_tolerance", "Unknown")),
                ("Early Exit Title:", parameters.get("early_exit_title", "Unknown")),
                ("Early Exit Author:", parameters.get("early_exit_author", "Unknown")),
                ("Early Exit Publisher:", parameters.get("early_exit_publisher", "Unknown")),
            ]

            for param_name, param_value in param_info:
                ws[f"A{param_row}"] = param_name
                ws[f"B{param_row}"] = str(param_value)
                param_row += 1

        # Status breakdown with explanations
        status_row = param_row + 2
        ws[f"A{status_row}"] = "Status Breakdown:"
        ws[f"A{status_row}"].font = Font(bold=True)
        status_row += 1

        # Headers for status table
        ws[f"A{status_row}"] = "Status"
        ws[f"B{status_row}"] = "Count"
        ws[f"C{status_row}"] = "Percentage"
        ws[f"D{status_row}"] = "Explanation"

        # Style headers
        for col in ["A", "B", "C", "D"]:
            cell = ws[f"{col}{status_row}"]
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        status_row += 1

        # Status explanations mapping (same as CSV exporter)
        explanations = {
            "US_PRE_": "Published before copyright expiration year - public domain due to age",
            "US_REGISTERED_NOT_RENEWED": (
                "Registered 1923-1977 but not renewed - entered public domain"
            ),
            "US_RENEWED": "Registered and renewed - still under copyright protection",
            "US_NO_MATCH": "No registration or renewal found - copyright status uncertain",
            "FOREIGN_PRE_": (
                "Foreign work published before copyright expiration - likely public domain"
            ),
            "FOREIGN_RENEWED": "Foreign work with US renewal - may be under copyright",
            "FOREIGN_REGISTERED_NOT_RENEWED": "Foreign work registered but not renewed in US",
            "FOREIGN_NO_MATCH": "Foreign work with no US records - copyright status uncertain",
            "COUNTRY_UNKNOWN_RENEWED": "Unknown country with renewal - likely under copyright",
            "COUNTRY_UNKNOWN_REGISTERED_NOT_RENEWED": "Unknown country registered but not renewed",
            "COUNTRY_UNKNOWN_NO_MATCH": (
                "Country of origin unknown - cannot determine copyright pathway"
            ),
            "OUT_OF_DATA_RANGE": "Published after our data coverage ends (1991)",
        }

        status_counts = metadata.get("status_counts", {})
        total_records_val = metadata.get("total_records", 0)

        # Ensure we have proper types
        if not isinstance(status_counts, dict):
            return

        total_records = 0
        if isinstance(total_records_val, (int, float)):
            total_records = int(total_records_val)

        if total_records > 0:
            # Sort statuses for consistent output
            sorted_statuses = sorted(status_counts.keys())

            for status in sorted_statuses:
                count_val = status_counts[status]
                if not isinstance(count_val, (int, float)):
                    continue

                count = int(count_val)
                if count == 0:
                    continue

                percentage = (count / total_records * 100) if total_records > 0 else 0

                # Find explanation by pattern matching
                explanation = ""
                status_upper = status.upper()
                for pattern, desc in explanations.items():
                    if pattern in status_upper or status_upper.startswith(pattern):
                        explanation = desc
                        # Add year/country specifics if present
                        if "_PRE_" in status_upper:
                            parts = status_upper.split("_")
                            for part in parts:
                                if part.isdigit() and len(part) == 4:
                                    explanation = explanation.replace(
                                        "copyright expiration year", part
                                    )
                        if "FOREIGN_" in status_upper:
                            # Extract country code (last 3 letters)
                            parts = status_upper.split("_")
                            if len(parts[-1]) == 3 and parts[-1].isalpha():
                                explanation = f"{explanation} ({parts[-1]})"
                        break

                if not explanation:
                    explanation = "Status requires further analysis"

                # Write status row
                display_name = self._format_status_as_sheet_name(status)
                ws[f"A{status_row}"] = display_name
                ws[f"B{status_row}"] = count
                ws[f"C{status_row}"] = f"{percentage:.1f}%"
                ws[f"D{status_row}"] = explanation

                # Apply status color to the status cell
                status_color = self._get_default_status_color(status)
                ws[f"A{status_row}"].fill = PatternFill(
                    start_color=status_color, end_color=status_color, fill_type="solid"
                )

                status_row += 1

            # Add total row
            ws[f"A{status_row}"] = "Total"
            ws[f"A{status_row}"].font = Font(bold=True)
            ws[f"B{status_row}"] = total_records
            ws[f"B{status_row}"].font = Font(bold=True)
            ws[f"C{status_row}"] = "100.0%"
            ws[f"C{status_row}"].font = Font(bold=True)
            ws[f"D{status_row}"] = "Total records analyzed"
            ws[f"D{status_row}"].font = Font(bold=True)

        # Auto-size columns
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 60

    def _create_data_sheet(self, wb: Workbook, sheet_name: str, records: JSONList) -> None:
        """Create a data sheet with records"""
        ws = wb.create_sheet(sheet_name)

        # Define columns
        columns = [
            ("MARC_ID", 15),
            ("MARC_Title", 40),
            ("MARC_Author", 25),
            ("MARC_Publisher", 25),
            ("MARC_Year", 10),
            ("MARC_Country", 12),
            ("Status_Rule", 30),
            ("Registration_ID", 15),
            ("Registration_Score", 15),
            ("Registration_Title_Match", 20),
            ("Registration_Author_Match", 20),
            ("Registration_Publisher_Match", 20),
            ("Registration_Year_Diff", 15),
            ("Renewal_ID", 15),
            ("Renewal_Score", 15),
            ("Renewal_Title_Match", 20),
            ("Renewal_Author_Match", 20),
            ("Renewal_Publisher_Match", 20),
            ("Renewal_Year_Diff", 15),
            ("Match_Type", 15),
            ("Data_Issues", 30),
            ("LCCN", 15),
        ]

        # Write headers
        for col_num, (header, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER

            # Set column width
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = width

        # Write data rows
        for row_num, record in enumerate(records, 2):
            if isinstance(record, dict):
                self._write_data_row(ws, row_num, record)

        # Freeze the header row
        ws.freeze_panes = "A2"

    def _write_data_row(self, ws: Worksheet, row_num: int, record: dict[str, JSONType]) -> None:
        """Write a single data row"""
        marc = record.get("marc", {})
        matches = record.get("matches", {})
        analysis = record.get("analysis", {})

        # Extract basic MARC data
        marc_id = ""
        original = {}
        metadata = {}

        if isinstance(marc, dict):
            id_val = marc.get("id", "")
            if isinstance(id_val, str):
                marc_id = id_val
            original_data = marc.get("original", {})
            if isinstance(original_data, dict):
                original = original_data
            metadata_data = marc.get("metadata", {})
            if isinstance(metadata_data, dict):
                metadata = metadata_data

        # Extract match data
        reg_match = {}
        ren_match = {}

        if isinstance(matches, dict):
            reg_data = matches.get("registration", {})
            if isinstance(reg_data, dict):
                reg_match = reg_data
            ren_data = matches.get("renewal", {})
            if isinstance(ren_data, dict):
                ren_match = ren_data

        # Determine match type
        match_types = []
        if reg_match.get("found") and reg_match.get("match_type") == "lccn":
            match_types.append("lccn")
        if ren_match.get("found") and ren_match.get("match_type") == "lccn":
            match_types.append("lccn")
        if not match_types:
            if reg_match.get("found") or ren_match.get("found"):
                match_types.append("similarity")

        match_type = "lccn" if "lccn" in match_types else ("similarity" if match_types else "none")

        # Format data issues
        data_issues = []

        if isinstance(analysis, dict):
            generic_data = analysis.get("generic_title", {})
            if isinstance(generic_data, dict) and generic_data.get("detected"):
                data_issues.append("generic_title")

            data_completeness = analysis.get("data_completeness", [])
            if isinstance(data_completeness, list):
                data_issues.extend([x for x in data_completeness if isinstance(x, str)])

        # Build row data
        row_data = [
            marc_id,
            original.get("title", ""),
            original.get("author_245c") or original.get("author_1xx", ""),
            original.get("publisher", ""),
            original.get("year", ""),
            metadata.get("country_code", ""),
            analysis.get("status_rule", "") if isinstance(analysis, dict) else "",
            reg_match.get("id", "") if reg_match.get("found") else "",
            self._format_score(reg_match),
            self._format_match_ratio(reg_match, "title"),
            self._format_author_match(reg_match),
            self._format_publisher_match(reg_match),
            self._format_year_diff(
                reg_match, str(original.get("year", "")) if original.get("year") else None
            ),
            ren_match.get("id", "") if ren_match.get("found") else "",
            self._format_score(ren_match),
            self._format_match_ratio(ren_match, "title"),
            self._format_author_match(ren_match),
            self._format_publisher_match(ren_match),
            self._format_year_diff(
                ren_match, str(original.get("year", "")) if original.get("year") else None
            ),
            match_type,
            ",".join(data_issues) if data_issues else "",
            metadata.get("lccn", ""),
        ]

        # Write data
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = self.DATA_ALIGNMENT
            cell.border = self.BORDER

            # Apply status coloring (column 7)
            if col_num == 7 and value:  # Status column
                status_color = self._get_default_status_color(str(value))
                cell.fill = PatternFill(
                    start_color=status_color, end_color=status_color, fill_type="solid"
                )

            # Apply score coloring
            elif col_num in [9, 15] and value and value != "LCCN":  # Score columns
                try:
                    score_val = float(str(value).rstrip("%"))
                    if score_val >= 90:
                        cell.fill = self.HIGH_SCORE_FILL
                    elif score_val >= 70:
                        cell.fill = self.MEDIUM_SCORE_FILL
                    else:
                        cell.fill = self.LOW_SCORE_FILL
                except ValueError:
                    pass

    def _format_score(self, match_data: JSONDict) -> str:
        """Format match score"""
        if not match_data.get("found"):
            return ""

        if match_data.get("match_type") == "lccn":
            return "LCCN"

        scores = match_data.get("scores", {})
        if isinstance(scores, dict):
            overall = scores.get("overall", 0)
            if isinstance(overall, (int, float)):
                return f"{overall:.0f}%"
        return "0%"

    def _format_match_ratio(self, match_data: JSONDict, field: str) -> str:
        """Format match ratio like '8/10'"""
        if not match_data.get("found"):
            return ""

        if match_data.get("match_type") == "lccn":
            return "exact"

        indicators = match_data.get("match_indicators", {})
        if isinstance(indicators, dict):
            field_data = indicators.get(field, {})
            if isinstance(field_data, dict):
                matched = field_data.get("words_matched", 0)
                total_val = field_data.get("words_total", 0)
                if isinstance(total_val, (int, float)) and total_val > 0:
                    return f"{matched}/{total_val}"

        return ""

    def _format_author_match(self, match_data: JSONDict) -> str:
        """Format author match type"""
        if not match_data.get("found"):
            return ""

        if match_data.get("match_type") == "lccn":
            return "exact"

        indicators = match_data.get("match_indicators", {})
        if isinstance(indicators, dict):
            author_data = indicators.get("author", {})
            if isinstance(author_data, dict):
                parts = []
                if author_data.get("surname_match"):
                    parts.append("surname")
                if author_data.get("given_match"):
                    parts.append("given")

                if parts:
                    return "/".join(parts)

        # Check for partial match
        scores = match_data.get("scores", {})
        if isinstance(scores, dict):
            author_score = scores.get("author", 0)
            if isinstance(author_score, (int, float)) and author_score > 0:
                return "partial"

        return "none" if match_data.get("found") else ""

    def _format_publisher_match(self, match_data: JSONDict) -> str:
        """Format publisher match as Y/N"""
        if not match_data.get("found"):
            return ""

        scores = match_data.get("scores", {})
        if isinstance(scores, dict):
            publisher_score = scores.get("publisher", 0)
            if isinstance(publisher_score, (int, float)) and publisher_score > 0:
                return "Y"
        return "N"

    def _format_year_diff(self, match_data: JSONDict, marc_year: str | None) -> str:
        """Format year difference"""
        if not match_data.get("found") or not marc_year:
            return ""

        original = match_data.get("original", {})
        if isinstance(original, dict):
            match_year = original.get("date") or original.get("year")
        else:
            match_year = None

        if not match_year:
            return ""

        try:
            diff = int(str(match_year)) - int(marc_year)
            if diff == 0:
                return "0"
            elif diff > 0:
                return f"+{diff}"
            else:
                return str(diff)
        except (ValueError, TypeError):
            return ""
