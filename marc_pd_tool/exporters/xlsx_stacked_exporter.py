# marc_pd_tool/exporters/xlsx_stacked_json_exporter.py

"""Stacked XLSX export functionality that reads from JSON data"""

# Standard library imports

# Third party imports
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.worksheet.worksheet import Worksheet

# Local imports
from marc_pd_tool.exporters.base_exporter import BaseJSONExporter
from marc_pd_tool.utils.types import JSONList
from marc_pd_tool.utils.types import JSONType


class StackedXLSXExporter(BaseJSONExporter):
    """Export stacked comparison XLSX from JSON data

    Creates Excel files with vertical stacked format showing
    original vs normalized text for detailed comparison.
    """

    # Header styling
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Subheader styling
    SUBHEADER_FONT = Font(bold=True)
    SUBHEADER_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    # Score coloring
    HIGH_SCORE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    MEDIUM_SCORE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    LOW_SCORE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Borders
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Status colors
    STATUS_COLORS = {
        "PD_PRE_MIN_YEAR": "E8F5E9",
        "PD_US_NOT_RENEWED": "D5F4E6",
        "PD_US_REG_NO_RENEWAL": "D4E6F1",
        "PD_US_NO_REG_DATA": "E3F2FD",
        "UNKNOWN_US_NO_DATA": "FFF9C4",
        "IN_COPYRIGHT": "FADBD8",
        "IN_COPYRIGHT_US_RENEWED": "FFCDD2",
        "RESEARCH_US_STATUS": "FCF3CF",
        "RESEARCH_US_ONLY_PD": "E8DAEF",
        "COUNTRY_UNKNOWN": "E5E7E9",
    }

    # Tab name mapping
    STATUS_TAB_NAMES = {
        "PD_PRE_MIN_YEAR": "PD Pre Min Year",
        "PD_US_NOT_RENEWED": "PD US Not Renewed",
        "PD_US_REG_NO_RENEWAL": "PD US Reg No Renewal",
        "PD_US_NO_REG_DATA": "PD US No Reg Data",
        "UNKNOWN_US_NO_DATA": "Unknown US No Data",
        "IN_COPYRIGHT": "In Copyright",
        "IN_COPYRIGHT_US_RENEWED": "In Copyright US Renewed",
        "RESEARCH_US_STATUS": "Research US Status",
        "RESEARCH_US_ONLY_PD": "Research US Only PD",
        "COUNTRY_UNKNOWN": "Country Unknown",
    }

    def _format_status_as_sheet_name(self, status: str) -> str:
        """Format a status string as a valid Excel sheet name

        Args:
            status: The status string (e.g., "US_RENEWED", "FOREIGN_NO_MATCH_GBR")

        Returns:
            A formatted sheet name suitable for Excel
        """
        # Excel sheet names can't exceed 31 characters
        # Replace underscores with spaces and title case
        formatted = status.replace("_", " ").title()

        # Truncate if necessary (leave room for potential numbering)
        if len(formatted) > 28:
            formatted = formatted[:28] + "..."

        return formatted

    def export(self) -> None:
        """Export records to stacked XLSX file"""
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create summary sheet
        self._create_summary_sheet(wb)

        if self.single_file:
            # All records in one sheet
            records = self.get_records()
            sorted_records = self.sort_by_quality(records)
            self._create_stacked_sheet(wb, "All Records", sorted_records)
        else:
            # Group by status and create sheets
            by_status = self.group_by_status()

            # Sort records within each status group
            for status in by_status:
                by_status[status] = self.sort_by_quality(by_status[status])

            # Create a sheet for each status
            # Sort statuses to ensure consistent order
            statuses = sorted(by_status.keys())

            for status in statuses:
                if by_status[status]:
                    # Get sheet name from mapping or generate from status
                    sheet_name = self.STATUS_TAB_NAMES.get(
                        status, self._format_status_as_sheet_name(status)
                    )
                    self._create_stacked_sheet(wb, sheet_name, by_status[status])

        # Save the workbook
        wb.save(self.output_path)

    def _create_summary_sheet(self, wb: Workbook) -> None:
        """Create summary sheet with statistics"""
        ws = wb.create_sheet("Summary")

        # Title
        ws["A1"] = "MARC PD Tool Stacked Results"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:B1")

        # Processing info
        metadata = self.get_metadata()
        ws["A3"] = "Processing Date:"
        ws["B3"] = metadata.get("processing_date", "Unknown")

        ws["A4"] = "Total Records:"
        ws["B4"] = metadata.get("total_records", 0)

        ws["A5"] = "Tool Version:"
        ws["B5"] = metadata.get("tool_version", "Unknown")

        # Status counts
        ws["A7"] = "Status Breakdown:"
        ws["A7"].font = Font(bold=True)

        row = 8
        status_counts = metadata.get("status_counts", {})
        if isinstance(status_counts, dict):
            for status, count in status_counts.items():
                display_name = self.STATUS_TAB_NAMES.get(status, status)
                ws[f"A{row}"] = display_name
                ws[f"B{row}"] = count
                row += 1

        # Auto-size columns
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _create_stacked_sheet(self, wb: Workbook, sheet_name: str, records: JSONList) -> None:
        """Create a sheet with stacked comparison format"""
        ws = wb.create_sheet(sheet_name)

        # Headers
        headers = [
            "Source",
            "ID",
            "Version",
            "Title",
            "Score",
            "Author",
            "Score",
            "Publisher",
            "Score",
            "Year",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        # Process records
        current_row = 2

        for record_idx, record in enumerate(records, 1):
            if isinstance(record, dict):
                # Add record header
                current_row = self._add_record_header(ws, current_row, record_idx, record)

                # Add MARC rows
                current_row = self._add_marc_rows(ws, current_row, record)

                # Add registration match
                current_row = self._add_match_row(
                    ws, current_row, record, "registration", "Registration"
                )

                # Add renewal match
                current_row = self._add_match_row(ws, current_row, record, "renewal", "Renewal")

                # Add spacing
                current_row += 1

        # Auto-size columns
        self._autosize_columns(ws)

    def _add_record_header(
        self, ws: Worksheet, row: int, record_num: int, record: dict[str, JSONType]
    ) -> int:
        """Add record header row"""
        marc = record.get("marc", {})
        analysis = record.get("analysis", {})

        marc_id = "Unknown"
        metadata: dict[str, JSONType] = {}
        if isinstance(marc, dict):
            id_val = marc.get("id", "Unknown")
            if isinstance(id_val, str):
                marc_id = id_val
            meta_data = marc.get("metadata", {})
            if isinstance(meta_data, dict):
                metadata = meta_data

        status = "UNKNOWN"
        if isinstance(analysis, dict):
            status_val = analysis.get("status", "UNKNOWN")
            if isinstance(status_val, str):
                status = status_val

        country = metadata.get("country_code", "")

        # Format header text
        header_text = f"Record {record_num} - ID: {marc_id} - Status: {status}"
        if country:
            header_text += f" - Country: {country}"

        # Merge cells for header
        ws.merge_cells(f"A{row}:J{row}")
        cell = ws[f"A{row}"]
        cell.value = header_text
        cell.font = self.SUBHEADER_FONT
        cell.fill = self.SUBHEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")

        # Add warnings if any
        warnings = ""
        if isinstance(analysis, dict):
            warnings = self._get_warnings(analysis)
        if warnings:
            row += 1
            ws.merge_cells(f"A{row}:J{row}")
            cell = ws[f"A{row}"]
            cell.value = f"⚠️ {warnings}"
            cell.font = Font(italic=True, color="FF6600")

        return row + 1

    def _add_marc_rows(self, ws: Worksheet, row: int, record: dict[str, JSONType]) -> int:
        """Add MARC original and normalized rows"""
        marc = record.get("marc", {})
        original: dict[str, JSONType] = {}
        normalized: dict[str, JSONType] = {}
        marc_id = ""

        if isinstance(marc, dict):
            id_val = marc.get("id", "")
            if isinstance(id_val, str):
                marc_id = id_val
            orig_data = marc.get("original", {})
            if isinstance(orig_data, dict):
                original = orig_data
            norm_data = marc.get("normalized", {})
            if isinstance(norm_data, dict):
                normalized = norm_data

        # Original row
        values = [
            "MARC",
            marc_id,
            "Original",
            original.get("title", ""),
            "-",
            original.get("author_245c") or original.get("author_1xx", ""),
            "-",
            original.get("publisher", ""),
            "-",
            original.get("year", ""),
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = self.THIN_BORDER

        row += 1

        # Normalized row
        values = [
            "MARC",
            marc_id,
            "Normalized",
            normalized.get("title", ""),
            "-",
            normalized.get("author", ""),
            "-",
            normalized.get("publisher", ""),
            "-",
            "-",
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = self.THIN_BORDER
            if col in [4, 6, 8]:  # Title, Author, Publisher columns
                cell.font = Font(name="Consolas", size=10)

        return row + 1

    def _add_match_row(
        self,
        ws: Worksheet,
        row: int,
        record: dict[str, JSONType],
        match_type: str,
        source_name: str,
    ) -> int:
        """Add a match row (registration or renewal)"""
        matches = record.get("matches", {})
        match_data: dict[str, JSONType] = {}
        if isinstance(matches, dict):
            data = matches.get(match_type, {})
            if isinstance(data, dict):
                match_data = data

        if not match_data.get("found"):
            # No match row
            values = [source_name, "-", "No match", "-", "-", "-", "-", "-", "-", "-"]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.THIN_BORDER
                cell.font = Font(color="999999")
            return row + 1

        # Match found
        original: dict[str, JSONType] = {}
        scores: dict[str, JSONType] = {}
        orig_data = match_data.get("original", {})
        if isinstance(orig_data, dict):
            original = orig_data
        score_data = match_data.get("scores", {})
        if isinstance(score_data, dict):
            scores = score_data

        match_method_val = match_data.get("match_type", "similarity")
        match_method = match_method_val if isinstance(match_method_val, str) else "similarity"

        # Format scores
        def format_score(score_value: float | None, is_lccn: bool = False) -> str:
            if is_lccn:
                return "LCCN"
            if score_value is None:
                return "-"
            return f"{score_value:.0f}%"

        is_lccn = match_method == "lccn"

        def get_score_value(field: str) -> float | None:
            val = scores.get(field)
            if isinstance(val, (int, float)):
                return float(val)
            return None

        match_id = ""
        id_val = match_data.get("id", "")
        if isinstance(id_val, str):
            match_id = id_val

        values = [
            source_name,
            match_id,
            "Original",
            str(original.get("title", "")),
            format_score(get_score_value("title"), is_lccn),
            str(original.get("author", "")),
            format_score(get_score_value("author"), is_lccn),
            str(original.get("publisher", "")),
            format_score(get_score_value("publisher"), is_lccn),
            str(original.get("date", "")),
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = self.THIN_BORDER

            # Apply score coloring
            if col in [5, 7, 9] and not is_lccn and value != "-":
                try:
                    score = float(value.rstrip("%"))
                    if score >= 90:
                        cell.fill = self.HIGH_SCORE_FILL
                    elif score >= 70:
                        cell.fill = self.MEDIUM_SCORE_FILL
                    else:
                        cell.fill = self.LOW_SCORE_FILL
                except ValueError:
                    pass
            elif is_lccn and col in [5, 7, 9]:
                cell.font = Font(bold=True, color="008000")

        return row + 1

    def _get_warnings(self, analysis: dict[str, JSONType]) -> str:
        """Get warning indicators for the record"""
        warnings = []

        # Check for generic title
        generic_info = analysis.get("generic_title", {})
        if isinstance(generic_info, dict) and generic_info.get("detected"):
            warnings.append("Generic title detected")

        # Check for data completeness issues
        data_issues = analysis.get("data_completeness", [])
        if isinstance(data_issues, list):
            warnings.extend([x for x in data_issues if isinstance(x, str)])

        # Add status rule if available
        status_rule_val = analysis.get("status_rule", "")
        if isinstance(status_rule_val, str) and status_rule_val:
            warnings.append(f"Rule: {status_rule_val}")

        return " | ".join(warnings) if warnings else ""

    def _autosize_columns(self, ws: Worksheet) -> None:
        """Auto-size columns based on content"""
        column_widths = {
            "A": 12,  # Source
            "B": 15,  # ID
            "C": 12,  # Version
            "D": 40,  # Title
            "E": 8,  # Score
            "F": 25,  # Author
            "G": 8,  # Score
            "H": 20,  # Publisher
            "I": 8,  # Score
            "J": 8,  # Year
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
