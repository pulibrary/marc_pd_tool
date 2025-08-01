# marc_pd_tool/exporters/xlsx_exporter.py

"""Standard XLSX export functionality that reads from JSON data"""

# Standard library imports

# Third party imports
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Local imports
from marc_pd_tool.exporters.base_exporter import BaseJSONExporter
from marc_pd_tool.utils.types import JSONDict
from marc_pd_tool.utils.types import JSONList
from marc_pd_tool.utils.types import JSONType


class XLSXExporter(BaseJSONExporter):
    """Export standard XLSX files from JSON data

    Creates Excel files with separate tabs for each copyright status,
    showing records in a tabular format with match details.
    """

    # Header styling
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data styling
    DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Score coloring
    HIGH_SCORE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    MEDIUM_SCORE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    LOW_SCORE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Status colors
    STATUS_COLORS = {
        "PD_DATE_VERIFY": "D4E6F1",
        "PD_NO_RENEWAL": "D5F4E6",
        "IN_COPYRIGHT": "FADBD8",
        "RESEARCH_US_STATUS": "FCF3CF",
        "RESEARCH_US_ONLY_PD": "E8DAEF",
        "COUNTRY_UNKNOWN": "E5E7E9",
    }

    # Tab name mapping
    STATUS_TAB_NAMES = {
        "PD_DATE_VERIFY": "PD Date Verify",
        "PD_NO_RENEWAL": "PD No Renewal",
        "IN_COPYRIGHT": "In Copyright",
        "RESEARCH_US_STATUS": "Research US Status",
        "RESEARCH_US_ONLY_PD": "Research US Only PD",
        "COUNTRY_UNKNOWN": "Country Unknown",
    }

    def export(self) -> None:
        """Export records to XLSX file"""
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create summary sheet
        self._create_summary_sheet(wb)

        if self.single_file:
            # All records in one sheet
            records = self.get_records()
            sorted_records = self.sort_by_quality(records)
            self._create_data_sheet(wb, "All Records", sorted_records)
        else:
            # Group by status and create sheets
            by_status = self.group_by_status()

            # Sort records within each status group
            for status in by_status:
                by_status[status] = self.sort_by_quality(by_status[status])

            # Create a sheet for each status
            status_order = [
                "PD_NO_RENEWAL",
                "PD_DATE_VERIFY",
                "IN_COPYRIGHT",
                "RESEARCH_US_STATUS",
                "RESEARCH_US_ONLY_PD",
                "COUNTRY_UNKNOWN",
            ]

            for status in status_order:
                if status in by_status and by_status[status]:
                    sheet_name = self.STATUS_TAB_NAMES.get(status, status)
                    self._create_data_sheet(wb, sheet_name, by_status[status])

        # Save the workbook
        wb.save(self.output_path)

    def _create_summary_sheet(self, wb: Workbook) -> None:
        """Create summary sheet with statistics"""
        ws = wb.create_sheet("Summary")

        # Title
        ws["A1"] = "MARC PD Tool Analysis Results"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:B1")

        # Processing info
        metadata = self.get_metadata()
        ws["A3"] = "Processing Date:"
        ws["B3"] = metadata.get("processing_date", "Unknown")

        ws["A4"] = "Total Records:"
        ws["B4"] = metadata.get("total_records", 0)

        ws["A5"] = "Tool Version:"
        ws["B5"] = metadata.get("tool_version", "Unknown")

        # Status counts
        ws["A7"] = "Status Breakdown:"
        ws["A7"].font = Font(bold=True)

        row = 8
        status_counts = metadata.get("status_counts", {})
        if isinstance(status_counts, dict):
            for status, count in status_counts.items():
                display_name = self.STATUS_TAB_NAMES.get(status, status)
                ws[f"A{row}"] = display_name
                ws[f"B{row}"] = count
                row += 1

        # Auto-size columns
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _create_data_sheet(self, wb: Workbook, sheet_name: str, records: JSONList) -> None:
        """Create a data sheet with records"""
        ws = wb.create_sheet(sheet_name)

        # Define columns
        columns = [
            ("MARC_ID", 15),
            ("MARC_Title", 40),
            ("MARC_Author", 25),
            ("MARC_Publisher", 25),
            ("MARC_Year", 10),
            ("MARC_Country", 12),
            ("Status_Rule", 30),
            ("Registration_ID", 15),
            ("Registration_Score", 15),
            ("Registration_Title_Match", 20),
            ("Registration_Author_Match", 20),
            ("Registration_Publisher_Match", 20),
            ("Registration_Year_Diff", 15),
            ("Renewal_ID", 15),
            ("Renewal_Score", 15),
            ("Renewal_Title_Match", 20),
            ("Renewal_Author_Match", 20),
            ("Renewal_Publisher_Match", 20),
            ("Renewal_Year_Diff", 15),
            ("Match_Type", 15),
            ("Data_Issues", 30),
            ("LCCN", 15),
        ]

        # Write headers
        for col_num, (header, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER

            # Set column width
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = width

        # Write data rows
        for row_num, record in enumerate(records, 2):
            if isinstance(record, dict):
                self._write_data_row(ws, row_num, record)

        # Freeze the header row
        ws.freeze_panes = "A2"

    def _write_data_row(self, ws: Worksheet, row_num: int, record: dict[str, JSONType]) -> None:
        """Write a single data row"""
        marc = record.get("marc", {})
        matches = record.get("matches", {})
        analysis = record.get("analysis", {})

        # Extract basic MARC data
        marc_id = ""
        original = {}
        metadata = {}

        if isinstance(marc, dict):
            id_val = marc.get("id", "")
            if isinstance(id_val, str):
                marc_id = id_val
            original_data = marc.get("original", {})
            if isinstance(original_data, dict):
                original = original_data
            metadata_data = marc.get("metadata", {})
            if isinstance(metadata_data, dict):
                metadata = metadata_data

        # Extract match data
        reg_match = {}
        ren_match = {}

        if isinstance(matches, dict):
            reg_data = matches.get("registration", {})
            if isinstance(reg_data, dict):
                reg_match = reg_data
            ren_data = matches.get("renewal", {})
            if isinstance(ren_data, dict):
                ren_match = ren_data

        # Determine match type
        match_types = []
        if reg_match.get("found") and reg_match.get("match_type") == "lccn":
            match_types.append("lccn")
        if ren_match.get("found") and ren_match.get("match_type") == "lccn":
            match_types.append("lccn")
        if not match_types:
            if reg_match.get("found") or ren_match.get("found"):
                match_types.append("similarity")

        match_type = "lccn" if "lccn" in match_types else ("similarity" if match_types else "none")

        # Format data issues
        data_issues = []

        if isinstance(analysis, dict):
            generic_data = analysis.get("generic_title", {})
            if isinstance(generic_data, dict) and generic_data.get("detected"):
                data_issues.append("generic_title")

            data_completeness = analysis.get("data_completeness", [])
            if isinstance(data_completeness, list):
                data_issues.extend([x for x in data_completeness if isinstance(x, str)])

        # Build row data
        row_data = [
            marc_id,
            original.get("title", ""),
            original.get("author_245c") or original.get("author_1xx", ""),
            original.get("publisher", ""),
            original.get("year", ""),
            metadata.get("country_code", ""),
            analysis.get("status_rule", "") if isinstance(analysis, dict) else "",
            reg_match.get("id", "") if reg_match.get("found") else "",
            self._format_score(reg_match),
            self._format_match_ratio(reg_match, "title"),
            self._format_author_match(reg_match),
            self._format_publisher_match(reg_match),
            self._format_year_diff(
                reg_match, str(original.get("year", "")) if original.get("year") else None
            ),
            ren_match.get("id", "") if ren_match.get("found") else "",
            self._format_score(ren_match),
            self._format_match_ratio(ren_match, "title"),
            self._format_author_match(ren_match),
            self._format_publisher_match(ren_match),
            self._format_year_diff(
                ren_match, str(original.get("year", "")) if original.get("year") else None
            ),
            match_type,
            ",".join(data_issues) if data_issues else "",
            metadata.get("lccn", ""),
        ]

        # Write data
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = self.DATA_ALIGNMENT
            cell.border = self.BORDER

            # Apply score coloring
            if col_num in [9, 15] and value and value != "LCCN":  # Score columns
                try:
                    score_val = float(str(value).rstrip("%"))
                    if score_val >= 90:
                        cell.fill = self.HIGH_SCORE_FILL
                    elif score_val >= 70:
                        cell.fill = self.MEDIUM_SCORE_FILL
                    else:
                        cell.fill = self.LOW_SCORE_FILL
                except ValueError:
                    pass

    def _format_score(self, match_data: JSONDict) -> str:
        """Format match score"""
        if not match_data.get("found"):
            return ""

        if match_data.get("match_type") == "lccn":
            return "LCCN"

        scores = match_data.get("scores", {})
        if isinstance(scores, dict):
            overall = scores.get("overall", 0)
            if isinstance(overall, (int, float)):
                return f"{overall:.0f}%"
        return "0%"

    def _format_match_ratio(self, match_data: JSONDict, field: str) -> str:
        """Format match ratio like '8/10'"""
        if not match_data.get("found"):
            return ""

        if match_data.get("match_type") == "lccn":
            return "exact"

        indicators = match_data.get("match_indicators", {})
        if isinstance(indicators, dict):
            field_data = indicators.get(field, {})
            if isinstance(field_data, dict):
                matched = field_data.get("words_matched", 0)
                total_val = field_data.get("words_total", 0)
                if isinstance(total_val, (int, float)) and total_val > 0:
                    return f"{matched}/{total_val}"

        return ""

    def _format_author_match(self, match_data: JSONDict) -> str:
        """Format author match type"""
        if not match_data.get("found"):
            return ""

        if match_data.get("match_type") == "lccn":
            return "exact"

        indicators = match_data.get("match_indicators", {})
        if isinstance(indicators, dict):
            author_data = indicators.get("author", {})
            if isinstance(author_data, dict):
                parts = []
                if author_data.get("surname_match"):
                    parts.append("surname")
                if author_data.get("given_match"):
                    parts.append("given")

                if parts:
                    return "/".join(parts)

        # Check for partial match
        scores = match_data.get("scores", {})
        if isinstance(scores, dict):
            author_score = scores.get("author", 0)
            if isinstance(author_score, (int, float)) and author_score > 0:
                return "partial"

        return "none" if match_data.get("found") else ""

    def _format_publisher_match(self, match_data: JSONDict) -> str:
        """Format publisher match as Y/N"""
        if not match_data.get("found"):
            return ""

        scores = match_data.get("scores", {})
        if isinstance(scores, dict):
            publisher_score = scores.get("publisher", 0)
            if isinstance(publisher_score, (int, float)) and publisher_score > 0:
                return "Y"
        return "N"

    def _format_year_diff(self, match_data: JSONDict, marc_year: str | None) -> str:
        """Format year difference"""
        if not match_data.get("found") or not marc_year:
            return ""

        original = match_data.get("original", {})
        if isinstance(original, dict):
            match_year = original.get("date") or original.get("year")
        else:
            match_year = None

        if not match_year:
            return ""

        try:
            diff = int(str(match_year)) - int(marc_year)
            if diff == 0:
                return "0"
            elif diff > 0:
                return f"+{diff}"
            else:
                return str(diff)
        except (ValueError, TypeError):
            return ""
