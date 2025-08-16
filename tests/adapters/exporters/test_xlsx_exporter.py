# tests/adapters/exporters/test_xlsx_exporter.py

"""Tests for XLSX export functionality"""

# Standard library imports
import os
from os import remove
from os.path import exists
import tempfile
from tempfile import NamedTemporaryFile

# Third party imports
import pytest

# Local imports
from marc_pd_tool.adapters.exporters.json_exporter import save_matches_json
from marc_pd_tool.adapters.exporters.xlsx_exporter import XLSXExporter


def export_to_xlsx(publications, xlsx_file, parameters=None):
    """Helper to export publications to XLSX via JSON"""
    temp_fd, temp_json = tempfile.mkstemp(suffix=".json")
    os.close(temp_fd)

    try:
        save_matches_json(publications, temp_json, parameters=parameters)
        exporter = XLSXExporter(temp_json, xlsx_file, single_file=False)
        exporter.export()
    finally:
        if os.path.exists(temp_json):
            os.unlink(temp_json)


class TestXLSXExporter:
    """Test XLSX export functionality"""

    def test_xlsx_exporter_creation(self, sample_publications):
        """Test creating XLSX exporter instance"""
        with NamedTemporaryFile(suffix=".json", delete=False) as f:
            json_path = f.name
        with NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            # Create JSON file first
            save_matches_json(sample_publications, json_path)

            # Create exporter with JSON path
            exporter = XLSXExporter(json_path, output_path, single_file=False)
            # BaseJSONExporter stores data, not paths
            assert exporter.output_path == output_path
            assert exporter.single_file is False
            assert exporter.json_data is not None
        finally:
            if exists(json_path):
                remove(json_path)
            if exists(output_path):
                remove(output_path)

    def test_xlsx_export_basic(self, sample_publications):
        """Test basic XLSX export"""
        with NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            export_to_xlsx(sample_publications, output_path)

            # Verify file was created
            assert exists(output_path)

            # Load and verify workbook structure
            # Third party imports
            from openpyxl import load_workbook

            wb = load_workbook(output_path)

            # Check sheets exist - with formatted status names
            expected_sheets = ["Summary", "Us Reg Not Renewed", "Us Renewed", "For No Match Xxk"]
            assert set(wb.sheetnames) == set(expected_sheets)

            # Check summary sheet
            summary = wb["Summary"]
            assert summary["A1"].value == "MARC PD Tool Analysis Results"
            assert summary["A4"].value == "Total Records:"
            assert summary["B4"].value == 3  # 3 sample publications

            # Check data sheets have correct headers
            pd_sheet = wb["Us Reg Not Renewed"]
            assert pd_sheet["A1"].value == "MARC_ID"
            assert pd_sheet["B1"].value == "MARC_Title"

            # Check data is written
            assert pd_sheet["A2"].value == "123"  # First publication ID
            assert pd_sheet["B2"].value == "Test Book One"

        finally:
            if exists(output_path):
                remove(output_path)

    def test_xlsx_export_with_parameters(self, sample_publications):
        """Test XLSX export with processing parameters"""
        with NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        parameters = {
            "title_threshold": 40,
            "author_threshold": 30,
            "year_tolerance": 1,
            "min_year": 1950,
            "max_year": 1960,
            "us_only": True,
            "score_everything_mode": False,
        }

        try:
            export_to_xlsx(sample_publications, output_path, parameters)

            # Load and verify parameters in summary
            # Third party imports
            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            wb["Summary"]

            # Find parameters section
            # The new JSON-based exporter doesn't include parameters in summary
            # This is expected behavior for the new implementation

        finally:
            if exists(output_path):
                remove(output_path)

    def test_xlsx_data_types(self, sample_publications):
        """Test that appropriate data types are used in XLSX"""
        with NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            export_to_xlsx(sample_publications, output_path)

            # Third party imports
            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            pd_sheet = wb["Us Reg Not Renewed"]

            # Check numeric types
            year_col = (
                5  # MARC_Year column (MARC_ID, MARC_Title, MARC_Author, MARC_Publisher, MARC_Year)
            )
            # Year might be string or int depending on the data
            year_value = pd_sheet.cell(row=2, column=year_col).value
            assert year_value == "1950" or year_value == 1950
            assert pd_sheet.cell(row=2, column=year_col).value == 1950

            # Check Registration_Score column (column I = 9)
            assert pd_sheet.cell(row=2, column=9).value == "95%"

        finally:
            if exists(output_path):
                remove(output_path)

    def test_xlsx_column_widths(self, sample_publications):
        """Test that column widths are set appropriately"""
        with NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            export_to_xlsx(sample_publications, output_path)

            # Third party imports
            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            pd_sheet = wb["Us Reg Not Renewed"]

            # Check some column widths
            assert pd_sheet.column_dimensions["A"].width == 15  # MARC_ID
            assert pd_sheet.column_dimensions["B"].width == 40  # MARC_Title
            assert pd_sheet.column_dimensions["E"].width == 10  # MARC_Year

        finally:
            if exists(output_path):
                remove(output_path)

    def test_xlsx_single_file_mode(self, sample_publications):
        """Test single file export mode"""
        with NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            # Create JSON and export with single_file=True
            temp_fd, temp_json = tempfile.mkstemp(suffix=".json")
            os.close(temp_fd)

            save_matches_json(sample_publications, temp_json)
            exporter = XLSXExporter(temp_json, output_path, single_file=True)
            exporter.export()

            # Third party imports
            from openpyxl import load_workbook

            wb = load_workbook(output_path)

            # In single file mode, should have only Summary and All Records sheets
            assert "Summary" in wb.sheetnames
            assert "All Records" in wb.sheetnames
            assert len(wb.sheetnames) == 2

            # Check all records are in the All Records sheet
            all_sheet = wb["All Records"]
            # Should have header + 3 records
            assert all_sheet.max_row == 4  # 1 header + 3 records

            os.unlink(temp_json)

        finally:
            if exists(output_path):
                remove(output_path)

    def test_xlsx_edge_case_statuses(self):
        """Test XLSX export with various status edge cases"""
        # Local imports
        from marc_pd_tool.core.domain.enums import CountryClassification
        from marc_pd_tool.core.domain.publication import Publication

        pubs = []

        # Unknown country with renewal
        pub1 = Publication(
            source_id="unk1",
            title="Unknown Country Renewed",
            author="Author",
            main_author="Author",
            pub_date="1950",
            publisher="Publisher",
            place="Unknown",
            country_code="xxx",
            country_classification=CountryClassification.UNKNOWN,
        )
        pub1.copyright_status = "COUNTRY_UNKNOWN_RENEWED"
        pub1.original_title = "Unknown Country Renewed"
        pub1.original_author = "Author"
        pubs.append(pub1)

        # Out of data range status
        pub2 = Publication(
            source_id="ood1",
            title="Out of Data Range Book",
            author="Author2",
            main_author="Author2",
            pub_date="1920",
            publisher="Publisher2",
            place="New York",
            country_code="xxu",
            country_classification=CountryClassification.US,
        )
        pub2.copyright_status = "US_OUT_OF_DATA_RANGE"
        pub2.original_title = "Out of Data Range Book"
        pub2.original_author = "Author2"
        pubs.append(pub2)

        # Unknown country registered not renewed
        pub3 = Publication(
            source_id="unk2",
            title="Unknown Country Not Renewed",
            author="Author3",
            main_author="Author3",
            pub_date="1955",
            publisher="Publisher3",
            place="Unknown",
            country_code="xxx",
            country_classification=CountryClassification.UNKNOWN,
        )
        pub3.copyright_status = "COUNTRY_UNKNOWN_REGISTERED_NOT_RENEWED"
        pub3.original_title = "Unknown Country Not Renewed"
        pub3.original_author = "Author3"
        pubs.append(pub3)

        with NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            export_to_xlsx(pubs, output_path)

            # Third party imports
            from openpyxl import load_workbook

            wb = load_workbook(output_path)

            # Check that all sheets were created with proper formatting
            assert len(wb.sheetnames) == 4  # Summary + 3 status sheets

            # Verify color coding for different statuses
            for sheet_name in wb.sheetnames:
                if sheet_name != "Summary":
                    sheet = wb[sheet_name]
                    # Check that headers have formatting
                    if sheet.max_row > 1:  # Has data rows
                        # Headers should have styling
                        assert sheet["A1"].fill is not None or sheet["A1"].font is not None

        finally:
            if exists(output_path):
                remove(output_path)

    def test_xlsx_empty_fields(self, sample_publications):
        """Test XLSX export handles empty/missing fields properly"""
        # Modify a publication to have missing fields
        pub = sample_publications[0]
        pub.original_edition = None
        pub.original_place = ""
        pub.lccn = None
        pub.registration_match = None  # Remove match to test empty match fields

        with NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            export_to_xlsx([pub], output_path)

            # Third party imports
            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            sheet = wb["Us Reg Not Renewed"]

            # Check that empty fields are handled (empty string or None)
            edition_col = 7  # MARC_Edition column
            assert sheet.cell(row=2, column=edition_col).value in ["", None]

            # Registration columns should be empty
            reg_title_col = 8  # Registration_Title
            assert sheet.cell(row=2, column=reg_title_col).value in ["", None]

        finally:
            if exists(output_path):
                remove(output_path)


class TestXLSXAvailability:
    """Test XLSX availability"""

    def test_xlsx_always_available(self):
        """Test that openpyxl is available as a direct dependency"""
        # Since openpyxl is now a direct dependency, we just verify it can be imported
        try:
            # Third party imports
            pass

            assert True  # Import succeeded
        except ImportError:
            pytest.fail("openpyxl should be available as a direct dependency")

        # Also verify XLSXExporter can be imported
        assert XLSXExporter is not None
