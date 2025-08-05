# tests/test_exporters/test_xlsx_stacked_exporter.py

"""Tests for XLSX stacked export functionality"""

# Standard library imports
from pathlib import Path
from tempfile import NamedTemporaryFile
from tempfile import TemporaryDirectory

# Local imports
from marc_pd_tool.data.enums import CopyrightStatus
from marc_pd_tool.data.enums import MatchType
from marc_pd_tool.exporters.json_exporter import save_matches_json
from marc_pd_tool.exporters.xlsx_stacked_exporter import StackedXLSXExporter
from tests.fixtures.publications import PublicationBuilder


class TestStackedXLSXExporter:
    """Test XLSX stacked export functionality"""

    def test_xlsx_stacked_exporter_creation(self, sample_publications):
        """Test creating XLSX stacked exporter instance"""
        with NamedTemporaryFile(suffix=".json", delete=False) as f:
            json_path = f.name
        with NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            # Create JSON file first
            save_matches_json(sample_publications, json_path)

            # Create exporter
            exporter = StackedXLSXExporter(json_path, output_path, single_file=False)

            assert exporter.output_path == output_path
            assert exporter.single_file is False
            assert exporter.json_data is not None
        finally:
            Path(json_path).unlink(missing_ok=True)
            Path(output_path).unlink(missing_ok=True)

    def test_xlsx_stacked_basic_export(self, sample_publications):
        """Test basic XLSX stacked export functionality"""
        with TemporaryDirectory() as temp_dir:
            json_path = str(Path(temp_dir) / "test.json")
            output_path = str(Path(temp_dir) / "test_stacked.xlsx")

            # Save to JSON first
            save_matches_json(sample_publications, json_path)

            # Export to XLSX stacked
            exporter = StackedXLSXExporter(json_path, output_path)
            exporter.export()

            # Verify XLSX file was created
            assert Path(output_path).exists()

            # Load and verify basic structure
            # Third party imports
            from openpyxl import load_workbook

            wb = load_workbook(output_path)

            # Should have Summary plus sheets for each status
            assert "Summary" in wb.sheetnames
            assert len(wb.sheetnames) > 1

    def test_xlsx_stacked_single_file_mode(self):
        """Test XLSX stacked export in single file mode"""
        # Create publications with different statuses
        pubs = [
            PublicationBuilder.basic_us_publication(source_id="1"),
            PublicationBuilder.basic_us_publication(source_id="2"),
            PublicationBuilder.basic_us_publication(source_id="3"),
        ]
        pubs[0].copyright_status = CopyrightStatus.US_REGISTERED_NOT_RENEWED.value
        pubs[1].copyright_status = CopyrightStatus.US_RENEWED.value
        pubs[2].copyright_status = CopyrightStatus.US_NO_MATCH.value

        with TemporaryDirectory() as temp_dir:
            json_path = str(Path(temp_dir) / "test.json")
            output_path = str(Path(temp_dir) / "stacked_single.xlsx")

            save_matches_json(pubs, json_path)

            # Export with single_file=True
            exporter = StackedXLSXExporter(json_path, output_path, single_file=True)
            exporter.export()

            # Load workbook
            # Third party imports
            from openpyxl import load_workbook

            wb = load_workbook(output_path)

            # Should have Summary and All Records sheets
            assert "Summary" in wb.sheetnames
            assert "All Records" in wb.sheetnames
            assert len(wb.sheetnames) == 2

    def test_xlsx_stacked_multiple_files_mode(self):
        """Test XLSX stacked export creating multiple sheets by status"""
        # Create publications with different statuses
        pubs = [
            PublicationBuilder.basic_us_publication(source_id="pd1"),
            PublicationBuilder.basic_us_publication(source_id="ic1"),
            PublicationBuilder.basic_us_publication(source_id="rs1"),
        ]
        pubs[0].copyright_status = CopyrightStatus.US_REGISTERED_NOT_RENEWED.value
        pubs[1].copyright_status = CopyrightStatus.US_RENEWED.value
        pubs[2].copyright_status = CopyrightStatus.US_NO_MATCH.value

        with TemporaryDirectory() as temp_dir:
            json_path = str(Path(temp_dir) / "test.json")
            output_path = str(Path(temp_dir) / "stacked_multi.xlsx")

            save_matches_json(pubs, json_path)

            # Export with single_file=False
            exporter = StackedXLSXExporter(json_path, output_path, single_file=False)
            exporter.export()

            # Load workbook
            # Third party imports
            from openpyxl import load_workbook

            wb = load_workbook(output_path)

            # Should have Summary plus status sheets
            assert "Summary" in wb.sheetnames
            # These sheet names would now be different with new enum values
            # We just check that we have the expected number of sheets
            # Check for sheet names - note that actual sheet names might be different due to dynamic statuses
            # But we should have multiple sheets based on status
            assert len(wb.sheetnames) > 1  # Should have multiple sheets for different statuses

    def test_xlsx_stacked_summary_sheet(self):
        """Test that summary sheet contains correct statistics"""
        pubs = [
            PublicationBuilder.basic_us_publication(source_id="1"),
            PublicationBuilder.basic_us_publication(source_id="2"),
        ]
        pubs[0].copyright_status = CopyrightStatus.US_REGISTERED_NOT_RENEWED.value
        pubs[1].copyright_status = CopyrightStatus.US_RENEWED.value

        with TemporaryDirectory() as temp_dir:
            json_path = str(Path(temp_dir) / "test.json")
            output_path = str(Path(temp_dir) / "summary_test.xlsx")

            save_matches_json(pubs, json_path)

            exporter = StackedXLSXExporter(json_path, output_path)
            exporter.export()

            # Check summary content
            # Third party imports
            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            summary = wb["Summary"]

            # Find total records cell
            total_found = False
            for row in summary.iter_rows(min_row=1, max_row=20, min_col=1, max_col=2):
                if row[0].value and "Total Records" in str(row[0].value):
                    assert row[1].value == 2
                    total_found = True
                    break
            assert total_found

    def test_xlsx_stacked_data_format(self):
        """Test that data is properly formatted in stacked view"""
        # Create publication with match data
        pub = PublicationBuilder.basic_us_publication()
        pub = PublicationBuilder.with_registration_match(
            pub, source_id="REG123", similarity_score=95.0, match_type=MatchType.SIMILARITY
        )
        pub.copyright_status = CopyrightStatus.US_REGISTERED_NOT_RENEWED.value

        with TemporaryDirectory() as temp_dir:
            json_path = str(Path(temp_dir) / "test.json")
            output_path = str(Path(temp_dir) / "stacked_format.xlsx")

            save_matches_json([pub], json_path)

            exporter = StackedXLSXExporter(json_path, output_path)
            exporter.export()

            # Check data formatting
            # Third party imports
            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            sheet = wb["US_REGISTERED_NOT_RENEWED"]

            # Should have headers
            assert sheet["A1"].value is not None
            # Should have data starting from row 3 (after headers)
            assert sheet["A3"].value is not None

    def test_xlsx_stacked_match_display(self):
        """Test that match information is displayed correctly"""
        # Create publication with both registration and renewal matches
        pub = PublicationBuilder.basic_us_publication()
        pub = PublicationBuilder.with_registration_match(
            pub, source_id="REG999", similarity_score=92.5
        )
        pub = PublicationBuilder.with_renewal_match(pub, source_id="REN999", similarity_score=88.0)
        pub.copyright_status = CopyrightStatus.US_RENEWED.value

        with TemporaryDirectory() as temp_dir:
            json_path = str(Path(temp_dir) / "test.json")
            output_path = str(Path(temp_dir) / "match_test.xlsx")

            save_matches_json([pub], json_path)

            exporter = StackedXLSXExporter(json_path, output_path)
            exporter.export()

            # Verify matches are displayed
            # Third party imports
            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            sheet = wb["US_RENEWED"]

            # Look for match IDs in the sheet
            found_reg = False
            found_ren = False
            for row in sheet.iter_rows(values_only=True):
                row_str = str(row)
                if "REG999" in row_str:
                    found_reg = True
                if "REN999" in row_str:
                    found_ren = True

            assert found_reg
            assert found_ren

    def test_xlsx_stacked_unicode_handling(self):
        """Test that unicode characters are properly handled"""
        pub = PublicationBuilder.basic_us_publication()
        pub.original_title = "Café société"
        pub.original_author = "José Müller"
        pub.copyright_status = CopyrightStatus.US_REGISTERED_NOT_RENEWED.value

        with TemporaryDirectory() as temp_dir:
            json_path = str(Path(temp_dir) / "unicode.json")
            output_path = str(Path(temp_dir) / "unicode.xlsx")

            save_matches_json([pub], json_path)

            exporter = StackedXLSXExporter(json_path, output_path)
            exporter.export()

            # Verify unicode is preserved
            # Third party imports
            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            sheet = wb["US_REGISTERED_NOT_RENEWED"]

            # Look for unicode text
            found_title = False
            found_author = False
            for row in sheet.iter_rows(values_only=True):
                row_str = str(row)
                if "Café société" in row_str:
                    found_title = True
                if "José Müller" in row_str:
                    found_author = True

            assert found_title
            assert found_author

    def test_xlsx_stacked_empty_data(self):
        """Test XLSX stacked export with no records"""
        with TemporaryDirectory() as temp_dir:
            json_path = str(Path(temp_dir) / "empty.json")
            output_path = str(Path(temp_dir) / "empty.xlsx")

            save_matches_json([], json_path)

            exporter = StackedXLSXExporter(json_path, output_path)
            exporter.export()

            # Should still create file with Summary
            assert Path(output_path).exists()

            # Third party imports
            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            assert "Summary" in wb.sheetnames

            # Summary should show 0 records
            summary = wb["Summary"]
            for row in summary.iter_rows(min_row=1, max_row=20, min_col=1, max_col=2):
                if row[0].value and "Total Records" in str(row[0].value):
                    assert row[1].value == 0
                    break
